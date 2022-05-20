import json
from collections import defaultdict
from dataclasses import dataclass
from os import listdir
from os.path import join
from typing import List

from matplotlib import pyplot as plt
from openpyxl import load_workbook

from models import DNA, EnhancedGenerationResult
from models.algo_result import TurnResult
from utils.plot import plot_multiple_best_scores_evolution_on_graph, plot_multiple_paretos_on_graph


@dataclass
class ExtractedResult:
    results: List[TurnResult]
    best_dna: DNA
    best_fitness: float
    generation: EnhancedGenerationResult
    selection: str
    mutation: str
    crossover: str
    file_name: str
    mutation_rate: float

    @property
    def name(self):
        return f'{self.selection}-{self.mutation}-{self.crossover}-{self.mutation_rate}'


def _load_data(folder: str = '../results') -> List[ExtractedResult]:
    extracted_results = []

    for c, file_name in enumerate(listdir(folder)):
        if not c % 100:
            print(c)
        if '.xlsx' in file_name:
            file_path = join(folder, file_name)

            try:
                wb = load_workbook(filename=file_path)

                general_sheet = wb['Général']
                best_fitness = general_sheet['B6'].value
                selection = general_sheet['B1'].value
                mutation = general_sheet['B2'].value
                crossover = general_sheet['B3'].value
                best_dna = _dna_from_string(general_sheet['B9'].value)
                mutation_rate = general_sheet['B10'].value

                results_sheet = wb["Results"]
                results: List[TurnResult] = []

                for i in range(2, results_sheet.max_row + 1):
                    results.append(TurnResult(
                        dna=_dna_from_string(results_sheet[f'A{i}'].value),
                        fitness=results_sheet[f'B{i}'].value,
                        distance_fitness=results_sheet[f'C{i}'].value,
                        risk_fitness=results_sheet[f'D{i}'].value,
                        generation=results_sheet[f'E{i}'].value,
                        boost_improvement=results_sheet[f'F{i}'].value,
                    ))

                with open(file_path.replace('.xlsx', '.json')) as json_file:
                    data = json.loads(json_file.read())

                enhanced_generation = {
                    tuple(a[0]): a[1] for a in data
                }

                extracted_results.append(ExtractedResult(
                    best_dna=best_dna,
                    best_fitness=best_fitness,
                    selection=selection,
                    mutation=mutation,
                    file_name=file_name.split('.')[0],
                    crossover=crossover,
                    results=results,
                    generation=enhanced_generation,
                    mutation_rate=mutation_rate
                ))
            except (FileNotFoundError, PermissionError):
                pass

    return extracted_results


def _dna_from_string(dna_string: str) -> DNA:
    return tuple(map(int, dna_string[1:-1].split(',')))


if __name__ == '__main__':
    data = _load_data()

    sorted_data = list(sorted(data, key=lambda x: x.best_fitness + x.results[-1].generation/1000000))

    for i in sorted_data:
        print(i.name)

    reduced_bests = sorted_data
    print(reduced_bests[0].file_name)
    plt.figure(figsize=(15, 15))
    plot_multiple_paretos_on_graph(plt, {
        item.name: item.generation
        for item in reduced_bests[0:20]
    })
    plt.show()

    plt.figure(figsize=(15, 15))
    plot_multiple_best_scores_evolution_on_graph(plt, {
        item.name: item.results
        for item in reduced_bests[0:100]
    })
    plt.show()

    a = []

    for i in data:
        a.append(i.name)

    results_dict = defaultdict(lambda: [])

    for item in data:
        results_dict[item.name].append(item.best_fitness, )

    avg_results_dict = {}

    for name, values in results_dict.items():
        avg_results_dict[name] = sum(values) / len(values)

    sorted_avg_results = sorted(avg_results_dict.items(), key=lambda x: x[1])

    for rank, i in enumerate(sorted_avg_results, start=1):
        print(f'#{rank},{i[0]},{i[1]}')

    print("Best algo", sorted_data[0].name)
    print("Best algo (avg)", sorted_avg_results[0])

    print(f"Algos testés: {len(set(a))}, Tests Effectués: {len(a)})")

    print("crossover,mutation,selection,fitness,generation")
    for i in sorted_data:
        print(f"{i.crossover},{i.mutation},{i.selection},{i.best_fitness},{i.results[-1].generation}")
